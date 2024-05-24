import pandas as pd
from openpyxl import load_workbook
import xlrd
from abc import ABC
from typing import Iterator
from langchain.docstore.document import Document
from langchain.document_loaders.base import BaseBlobParser
from langchain.document_loaders.blob_loaders.schema import Blob
import warnings
from cat.log import log
import datetime

class XLSXExcelParser(BaseBlobParser, ABC):
    def lazy_parse(self, blob: Blob) -> Iterator[Document]:
        warnings.simplefilter(action='ignore', category=UserWarning)
        with blob.as_bytes_io() as file:
            if blob.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                wb = load_workbook(file, data_only=True)
                log.info(wb)
            for page in wb.sheetnames:
                df = pd.read_excel(file, sheet_name=page, keep_default_na=False,engine='openpyxl')
                df.dropna(axis=0, how='all', inplace=True)  
                df.dropna(axis=1, how='all', inplace=True)  
                sheet_data_list = df.to_dict(orient='records')  
                filtered_data_list = [
                        {key: value if not isinstance(value, str) else value
                         for key, value in record.items() if value != ''}
                        for record in sheet_data_list
                    ]
                filtered_data_list = [record for record in filtered_data_list if record]  
                for record in filtered_data_list:
                    for key, value in record.items():
                        if isinstance(value, datetime.datetime):
                            record[key] = value.strftime('%d %m %Y %h, %H:%M:%S')
                    yield Document(page_content=str(record).replace(r"\n"," "), metadata={})

class XLSExcelParser(BaseBlobParser, ABC):
    def lazy_parse(self, blob: Blob) -> Iterator[Document]:
        warnings.simplefilter(action='ignore', category=UserWarning)
        with blob.as_bytes_io() as file:
            if blob.mimetype == "application/vnd.ms-excel":
                wb = pd.read_excel(file, sheet_name=None, engine='xlrd')
            for sheet_name, sheet_data in wb.items():
                sheet_data_list = sheet_data.to_dict(orient='records')
                for record in sheet_data_list:
                    yield Document(page_content=str(record), metadata={})

class CSVParser(BaseBlobParser):
    def lazy_parse(self, blob: Blob) -> Iterator[Document]:
        warnings.simplefilter(action='ignore', category=UserWarning)
        with blob.as_bytes_io() as file:
            if blob.mimetype == "text/csv":
                df = pd.read_csv(file)
                content = df.to_dict(orient='records')
                for record in content:
                    yield Document(page_content=str(record), metadata={})
